import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
import matplotlib.lines as mlines
import matplotlib.ticker as mtick
import sys
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()

# data
letters = string.ascii_uppercase
benchmark_no = 2
scenarios = ['5-5','10-10','15-10']
titles = ['(a) n = 5, m = 5', '(b) n = 10, m = 10', '(c) n = 15, m = 10']
runs = 10

# start with a square Figure
fig = plt.figure(figsize=(12,5))
all_scenarios = gridspec.GridSpec(1, 3, figure=fig)

for i in range(3):
    # data
    path = sys.path[0]+'\\Experiment_result.xlsx'
    data = load_workbook(path)[scenarios[i]]
    # plots
    section = gridspec.GridSpecFromSubplotSpec(100, 1000, subplot_spec = all_scenarios[i])
    ''' the cumulative tardiness part'''
    ax1 = fig.add_subplot(section[:85,:500])
    ax1.invert_xaxis()
    ax1.set_xlim(left=0.9, right=0)
    ax1.set_xticks(np.arange(1,0,-0.1))
    plt.setp(ax1.get_xticklabels(), rotation=45, ha='right', rotation_mode="anchor", fontsize=8.5)
    ax1.xaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    ax1.set_ylim(bottom=0, top=runs+1)
    ax1.set_yticks(np.arange(1,runs+1))
    ax1.set_title(titles[i], ha='left', fontsize=12)
    ax1.set_xlabel('Performance Gain %', fontsize=10)
    if i == 0:
        ax1.set_ylabel('Index of problem instance', fontsize=11)
    # retrive and plot the NCT data
    sum = []
    for idx in range(3):
        sum.append([data[letters[idx+1] + str(i)].value for i in range(3,3+runs)])
    print(sum)
    sum = 1 - sum / np.array(sum[0])
    sum = np.delete(sum,0,axis=0)
    print(sum)
    ax1.barh(np.arange(1,runs+1), sum[0], height=0.3, align='edge', ec = 'k', zorder=3, color = 'darkgray')
    ax1.barh(np.arange(1,runs+1)-0.3, sum[1], height=0.3, align='edge', ec = 'k', zorder=3, color = '#00ffff')
    ax1.grid()


    ''' the number of tardy job part '''
    ax2 = fig.add_subplot(section[:85,500:])
    ax2.get_yaxis().set_visible(False)
    ax2.set_xlim(left=0, right=(i//2+1)*5)
    ax2.set_xticks(np.arange(-1, (i//2+1)*5)+1)
    #plt.setp(ax2.get_xticklabels(), rotation=-45, ha='right', rotation_mode="anchor", fontsize=8.5)
    ax2.set_ylim(bottom=0, top=runs+1)
    ax2.set_xlabel('No. of tardy jobs', fontsize=10)
    # retrive and plot the no. of tardy job data
    no = []
    for idx in range(3):
        no.append([data[letters[idx+5] + str(i)].value for i in range(3,3+runs)])
    print(no)
    ax2.barh(np.arange(1,runs+1), no[0], height=0.3, align='edge', ec = 'k', zorder=3, color = 'darkgray')
    ax2.barh(np.arange(1,runs+1)-0.3, no[1], height=0.3, align='edge', ec = 'k', zorder=3, color = '#00ffff')
    ax2.grid()

    ax1.vlines(0,0,100, color='k')
    ax2.vlines(0,0,100, color='k')

section = gridspec.GridSpecFromSubplotSpec(100, 1000, subplot_spec = all_scenarios[1])
ax3 = fig.add_subplot(section[99:100,:])
# common legend
color_list = ['darkgray', '#00ffff']
name = ['GA', 'deep MARL-RS']
print(name)
legend_elements = [Patch(facecolor=color_list[i], edgecolor='k',label=name[i]) for i in range(2)]
ax3.legend(handles=legend_elements, fontsize=9, loc=9, ncol=2)
ax3.axis('off')


fig.subplots_adjust(top=0.95, bottom=0.15, wspace=0.1)
fig.savefig(sys.path[0]+"/static_result.png", dpi=500, bbox_inches='tight')
plt.show()
