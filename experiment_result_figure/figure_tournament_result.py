import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
import matplotlib.lines as mlines
import matplotlib.ticker as mtick
import sys
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()


# index to retrive the data
panel_A = ['(a.1) Normalized cumulative tardiness\nUtilization rate = 70%','(a.2) Normalized cumulative tardiness\nUtilization rate = 80%','(a.3) Normalized cumulative tardiness\nUtilization rate = 90%']
panel_B = ['(b.1) Win rate %\nUtilization rate = 70%','(b.2) Win rate %\nUtilization rate = 80%','(b.3) Win rate %\nUtilization rate = 90%']

letters = string.ascii_uppercase
benchmark_no = 8
scenarios = [70,80,90]
runs = 100
c_l = ['k','#d46a7e','#fe019a','#86a17d','#978a84','#b7c9e2','#95d0fc','#029386','#f97306','#c2b709','#c20078','#00035b','#75bbfd','#929591','#89fe05','#8f1402','#9a0eea','#033500','#06c2ac','#ffffe4','#8af1fe','#d1b26f','#00ffff','#13eac9']
pie_color_list = c_l[:benchmark_no-1]+['#8af1fe']
bplot_color_list = ['w' for i in range(benchmark_no-2)]+['#8af1fe']
'''plot'''
# gridspec inside gridspec
fig = plt.figure(figsize=(10,7),)
ax = {}
sec_ax = {}
all_scenarios = gridspec.GridSpec(1, 3, figure=fig)


for ax_idx in range(3):
    '''import the data'''
    path = sys.path[0]+'\\tour_result_{}.xlsx'.format(scenarios[ax_idx])
    data_sum = load_workbook(path)['sum']
    data_win_rate = load_workbook(path)['win rate']
    '''create the grids'''
    section = gridspec.GridSpecFromSubplotSpec(20, 20, subplot_spec = all_scenarios[ax_idx])
    ax[ax_idx] = fig.add_subplot(section[1:9, :])
    # set different range to create broken axis
    ax[ax_idx].set_ylim(bottom=0, top=0.9)
    ax[ax_idx].set_xlim(left = -1, right=benchmark_no-1)
    '''retrive the data'''
    name = []
    sum = []
    # retrive the data
    for idx in range(benchmark_no):
        name.append(data_sum[letters[idx] + '1'].value)
        sum.append([data_sum[letters[idx] + str(i)].value for i in range(2,2+runs)])
    #print(name)
    #print(sum)
    sum = 1 - sum / np.array(sum[0])
    name.pop(0) # drop the I-DQN
    sum = np.delete(sum,0,axis=0)
    name[-1] = r'$\mathbf{deep}$'+' '+ r'$\mathbf{MARL-RS}$'
    # create the plots
    x_position = np.arange(len(name))
    '''plot the data'''
    bplot = ax[ax_idx].boxplot(sum.transpose(), positions=x_position, showmeans=True, meanline=True, patch_artist=True, notch=True, zorder=3,)
    for patch, c in zip(bplot['boxes'], bplot_color_list):
        patch.set_facecolor(c)
    # ax[ax_idx].violinplot(sum.transpose(), positions=x_position, showmeans=True, )
    # ticks
    ax[ax_idx].set_yticks(np.arange(-0.2,1,0.1))
    ax[ax_idx].yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    ax[ax_idx].set_xticks(x_position)
    ax[ax_idx].set_xticklabels(name)
    #ax[ax_idx].set_yticks(np.arange(0, 1.5, 0.1))
    plt.setp(ax[ax_idx].get_xticklabels(), rotation=27, ha='right', rotation_mode="anchor", fontsize=9)
    plt.setp(ax[ax_idx].get_yticklabels(), fontsize=9)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )
    # lines
    ax[ax_idx].hlines(y=0, xmin=-1, xmax=benchmark_no, colors='k', linestyles='solid', linewidths=1.5)
    ax[ax_idx].hlines(y=sum[-1].mean(), xmin=-1, xmax=benchmark_no, colors='g', linestyles='--', linewidths=1.5, zorder=1)
    # label
    ax[ax_idx].set_title(panel_A[ax_idx], fontsize=10)
    legend_color = ['k','g']
    legend_line = ['-','--']
    legend_label = ['I-DQN baseline','deep MARL-RS mean']
    legend_elements = [mlines.Line2D([], [], color=legend_color[i], linestyle=legend_line[i], markersize=5, label=legend_label[i]) for i in range(2)]
    ax[ax_idx].legend(handles=legend_elements, fontsize=8, loc=2, ncol=1)



    '''common legend'''
    if ax_idx == 1:
        ax[200] = fig.add_subplot(section[19:, :])
        ec_c = ['w']*7 + ['k']
        # common legend
        name.insert(0,'I-DQN*')
        legend_elements = [Patch(facecolor=pie_color_list[i], edgecolor = ec_c[i], label=name[i]) for i in range(benchmark_no)]
        ax[200].legend(handles=legend_elements, fontsize=9, loc='center', ncol=int(benchmark_no/2))
        ax[200].axis('off')

    '''and the pie chart'''
    ax[ax_idx+10] = fig.add_subplot(section[13:18, 4:17])
    '''retrive the data'''
    name = []
    win_rate = []
    colors = []
    for idx in range(benchmark_no):
        rate = data_win_rate[letters[idx] + '2'].value
        if rate>0 :
            name.append(data_win_rate[letters[idx] + '1'].value)
            win_rate.append(rate)
            colors.append(pie_color_list[idx])
    '''labelling'''
    no = len(win_rate)
    explode = [0 for i in range(len(name)-1)]+[0.15]  # only "explode" DRL
    label = [None if win_rate[i]<=2 and win_rate[i-1]<=3  else str(win_rate[i]) for i in range(no)]
    annotates = [str(win_rate[i]) if win_rate[i]<=2 and win_rate[i-1]<=3 else None for i in range(no)]
    wedges,texts = ax[ax_idx+10].pie(win_rate, labels = label, explode=explode, labeldistance=1.2, colors=colors, startangle=180, wedgeprops=dict(ec='w'), textprops=dict(fontsize=8.5))
    plt.setp(texts, ha='center', )
    plt.setp(texts[-1], fontsize=10.5, fontweight='bold')
    plt.setp(wedges[-1], ec='k')
    kw = dict(arrowprops=dict(arrowstyle="-",lw=0.5),  va="center", fontsize=8.5, )
    pre=1
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        if annotates[i]:
            if np.abs(pre-y) > 0.15 or pre==1:
                ax[ax_idx+10].annotate(annotates[i], xy=(x, y), xytext=(1.35*np.sign(x), 1.4*y),horizontalalignment=horizontalalignment, **kw)
            else:
                ax[ax_idx+10].annotate(annotates[i], xy=(x, y), xytext=(1.35*np.sign(x), 1.4*pre-0.18),horizontalalignment=horizontalalignment, **kw)
            pre=y
    ax[ax_idx+10].axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    ax[ax_idx+10].set_title(panel_B[ax_idx], fontsize=10)

ax[0].set_ylabel('Performance Gain %', fontsize=10)


fig.subplots_adjust(top=0.9, bottom=0.1, hspace=0.5)
fig.savefig(sys.path[0]+"/pics/tournament_result.png", dpi=600, bbox_inches='tight')
plt.show()
