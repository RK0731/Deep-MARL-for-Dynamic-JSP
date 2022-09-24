import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
import matplotlib.lines as mlines
import matplotlib.ticker as mtick
import sys
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()


# index to retrive the data
panel = ['(a) Tardy rate, UTIL.=70%','(b) Tardy rate, UTIL.=80%','(c) Tardy rate, UTIL.=90%']

letters = string.ascii_uppercase
benchmark_no = 20
scenarios = [70,80,90]
runs = 100
base = ['#7e1e9c','#15b01a','#0343df','#ff81c0','#653700','#e50000','#95d0fc','#029386','#f97306','#c2b709','#c20078','#00035b','#75bbfd','#929591','#89fe05','#8f1402','#9a0eea','#033500','#06c2ac','#ffffe4','#25a36f','#b1916e']
addon = ['#8af1fe']
color_list = base[:benchmark_no]+addon
bplot_color_list = ['k'] + ['w' for i in range(benchmark_no-1)]+addon

'''plot'''
# gridspec inside gridspec
fig = plt.figure(figsize=(10,12))
ax = {}
bound =28

for ax_idx in range(3):
    '''import the data'''
    path = sys.path[0]+'\\exp_result_{}.xlsx'.format(scenarios[ax_idx])
    data = load_workbook(path)['tardy rate']
    ax[ax_idx] = fig.add_subplot(3,1,ax_idx+1)
    # set different range to create broken axis
    ax[ax_idx].set_ylim(bottom=0.1, top=0.9)
    ax[ax_idx].set_xlim(left = -1, right=benchmark_no+1)
    '''retrive the data'''
    name = []
    rate = []
    # retrive the data
    for idx in range(benchmark_no+1):
        name.append(data[letters[idx] + '1'].value)
        rate.append([data[letters[idx] + str(i)].value for i in range(2,2+runs)])
    rate = np.array(rate)
    # create the plots
    x_position = np.arange(len(name))
    '''plot the data'''
    bplot = ax[ax_idx].boxplot(rate.transpose(), positions=x_position, showmeans=True, meanline=True, patch_artist=True, notch=True, zorder=3,)
    for patch, c in zip(bplot['boxes'], bplot_color_list):
        patch.set_facecolor(c)
    # ax[ax_idx].violinplot(rate.transpose(), positions=x_position, showmeans=True, )
    # ticks
    ax[ax_idx].set_yticks(np.arange(0.1,1.1,0.1))
    ax[ax_idx].yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    ax[ax_idx].set_xticks(x_position)
    ax[ax_idx].set_xticklabels(name)
    ax[ax_idx].set_ylabel('Tardy rate %', fontsize=11)
    #ax[ax_idx].set_yticks(np.arange(0, 1.5, 0.1))
    plt.setp(ax[ax_idx].get_xticklabels(), rotation=25, ha='right', rotation_mode="anchor", fontsize=9)
    plt.setp(ax[ax_idx].get_yticklabels(), fontsize=9)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )
    # lines
    ax[ax_idx].hlines(y=rate[0].mean(), xmin=-1, xmax=benchmark_no+2, colors='k', linestyles='solid', linewidths=1.5)
    ax[ax_idx].hlines(y=rate[-1].mean(), xmin=-1, xmax=benchmark_no+2, colors='g', linestyles='--', linewidths=1, zorder=1)
    # fill
    ax[ax_idx].fill_between([-1,benchmark_no+1], [rate[0].mean(),rate[0].mean()], [0,0],alpha=0.2)
    # label
    ax[ax_idx].set_title(panel[ax_idx], fontsize=12)
    # legends
    legend_color = ['k','b','r']
    legend_line = ['-','--','--']
    legend_label = ['FIFO baseline','mean of deep MARL-AS','mean of deep MARL-MR']
    legend_elements = [mlines.Line2D([], [], color=legend_color[i], linestyle=legend_line[i], markersize=5, label=legend_label[i]) for i in range(3)]
    ax[ax_idx].legend(handles=legend_elements, fontsize=8, loc=2, ncol=3)


fig.subplots_adjust(top=0.95, bottom=0.1, hspace=0.5)
#fig.savefig(sys.path[0]+"/tardy_rate.png".format(scenarios[ax_idx]), dpi=500, bbox_inches='tight')
plt.show()
