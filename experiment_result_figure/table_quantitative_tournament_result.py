import sys
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
import pandas as pd
from pandas import DataFrame


letters = string.ascii_uppercase
comparison_pair = [[1,3],[2,7],[1,2],[3,7]]
scenarios = [70,80,90]
runs = 100

output_data = {}
title =[]

for idx,scenario in enumerate(scenarios):
    '''import the data'''
    path = sys.path[0]+'\\tour_result_{}.xlsx'.format(scenario)
    data_sum = load_workbook(path)['sum']
    '''retrive the data'''
    NPs = []
    title = []
    for pair in comparison_pair:
        name_base = data_sum[letters[pair[0]] + '1'].value
        sum_base = [data_sum[letters[pair[0]] + str(i)].value for i in range(2,2+runs)]
        name_target = data_sum[letters[pair[1]] + '1'].value
        sum_target = [data_sum[letters[pair[1]] + str(i)].value for i in range(2,2+runs)]
        #print(name_base)
        #print(sum_base)
        #print(name_target)
        #print(sum_target)
        NP = 1 - sum_target / np.array(sum_base)
        title.append('{} / {}'.format(name_base, name_target))
        NPs.append([NP.mean()])
        print(NPs)
        #print(title)
    output_data[idx] = DataFrame(np.transpose(NPs), columns=title)

address = sys.path[0]+'\\RAW_quantitative_tour.xlsx'
Excelwriter = pd.ExcelWriter(address,engine="xlsxwriter")
dflist = [output_data[idx] for idx,scenario in enumerate(scenarios)]
sheetname = ['70','80','90']

for i,df in enumerate(dflist):
    df.to_excel(Excelwriter, sheet_name=sheetname[i], index=False)
Excelwriter.save()
print('export to {}'.format(address))
