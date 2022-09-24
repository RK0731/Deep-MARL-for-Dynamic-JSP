import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
import matplotlib.lines as mlines
import matplotlib.ticker as mtick
import sys
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()


# index to retrive the data
panel_A = ['(a.1) Normalized cumulative tardiness, Utilization rate=70%','(a.2) Normalized cumulative tardiness, Utilization rate=80%','(a.3) Normalized cumulative tardiness, Utilization rate=90%']
panel_B = ['(b.1) Win rate %\nUtil. = 70%\ndeep MARL(✖)','(b.2) Win rate %\nUtil. = 80%\ndeep MARL(✖)','(b.3) Win rate %\nUtil. = 90%\ndeep MARL(✖)']
panel_C = ['(c.1) Win rate %\nUtil. = 70%\ndeep MARL(✔)','(c.2) Win rate %\nUtil. = 80%\ndeep MARL(✔)','(c.3) Win rate %\nUtil. = 90%\ndeep MARL(✔)']

letters = string.ascii_uppercase
benchmark_no = 20
scenarios = [70,80,90]
runs = 100
color_list = ['#7e1e9c','#15b01a','#0343df','#ff81c0','#653700','#e50000','#95d0fc','#029386','#f97306','#c2b709','#c20078','#00035b','#75bbfd','#929591','#89fe05','#8f1402','#9a0eea','#033500','#06c2ac','#ffffe4','#8af1fe','#d1b26f','#00ffff','#13eac9']
bplot_color_list = ['None' for i in range(benchmark_no-1)]+['#8af1fe']
'''plot'''
# gridspec inside gridspec
fig = plt.figure(figsize=(12,13))
ax = {}
all_scenarios = gridspec.GridSpec(1, 1, figure=fig)
section = gridspec.GridSpecFromSubplotSpec(32, 37, subplot_spec = all_scenarios[0])

for ax_idx in range(3):
    upper = ax_idx*10+1
    lower = ax_idx*10+8
    '''import the data'''
    path = sys.path[0]+'\\exp_result_{}.xlsx'.format(scenarios[ax_idx])
    data_sum = load_workbook(path)['sum']
    data_before_win_rate = load_workbook(path)['before win rate']
    data_win_rate = load_workbook(path)['win rate']
    '''create the grids'''
    ax[ax_idx] = fig.add_subplot(section[ax_idx*11:ax_idx*11+8, :21])
    # set different range to create broken axis
    ax[ax_idx].set_ylim(bottom=0, top=0.9)
    ax[ax_idx].set_xlim(left = -1, right=benchmark_no)

    '''retrive the data'''
    name = []
    sum = []
    # retrive the data
    for idx in range(benchmark_no+1):
        name.append(data_sum[letters[idx] + '1'].value)
        sum.append([data_sum[letters[idx] + str(i)].value for i in range(2,2+runs)])
    name = ['deep MARL-RS' if x=='deep MARL' else x for x in name]
    #print(sum)
    sum = 1 - sum / np.array(sum[0])

    ''' adjust the name'''
    name.pop(0) # drop the FIFO
    sum = np.delete(sum,0,axis=0)
    name[-1] = r'$\mathbf{deep}$'+' '+ r'$\mathbf{MARL-RS}$'
    x_position = np.arange(len(name))
    replace = ['CR+SPT','LWKR+SPT','LWKR+MOD','PT+WINQ','PT+WINQ+S','2PT+LWKR+S','2PT+WINQ+NPT']
    for i in range(12,19):
        name[i] = replace[i-12]
    #print(name)

    '''plot the data'''
    bplot = ax[ax_idx].boxplot(sum.transpose(), positions=x_position, showmeans=True, meanline=True, patch_artist=True, notch=True, zorder=3,)
    for patch, c in zip(bplot['boxes'], bplot_color_list):
        patch.set_facecolor(c)
    # ax[ax_idx].violinplot(sum.transpose(), positions=x_position, showmeans=True, )
    # ticks
    ax[ax_idx].set_yticks(np.arange(-0.2,1,0.1))
    ax[ax_idx].yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    ax[ax_idx].set_xticks(x_position)
    ax[ax_idx].set_xticklabels(name)
    ax[ax_idx].set_ylabel('Performance Gain %', fontsize=10)
    #ax[ax_idx].set_yticks(np.arange(0, 1.5, 0.1))
    plt.setp(ax[ax_idx].get_xticklabels(), rotation=35, ha='right', rotation_mode="anchor", fontsize=9)
    plt.setp(ax[ax_idx].get_yticklabels(), fontsize=9)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )
    # lines
    ax[ax_idx].hlines(y=0, xmin=-1, xmax=benchmark_no, colors='k', linestyles='solid', linewidths=1.5)
    ax[ax_idx].hlines(y=sum[-1].mean(), xmin=-1, xmax=benchmark_no, colors='g', linestyles='--', linewidths=1.5, zorder=1)
    # label
    ax[ax_idx].set_title(panel_A[ax_idx], fontsize=12)

    legend_color = ['k','g']
    legend_line = ['-','--']
    legend_label = ['FIFO baseline','deep MARL-RS mean']
    legend_elements = [mlines.Line2D([], [], color=legend_color[i], linestyle=legend_line[i], markersize=5, label=legend_label[i]) for i in range(2)]
    ax[ax_idx].legend(handles=legend_elements, fontsize=8, loc=2, ncol=2)



    ''' legend of pir charts'''
    if ax_idx == 2:
        ax[100] = fig.add_subplot(section[31:, 22:])
        ec_c = ['w']*(len(name)-1) + ['k']
        # common legend
        legend_elements = [Patch(facecolor=color_list[i+1], edgecolor = ec_c[i], label=name[i]) for i in range(benchmark_no)]
        ax[100].legend(handles=legend_elements, ncol=3, fontsize=9, loc=8)
        ax[100].axis('off')



    '''and the pie chart'''
    ax[ax_idx+10] = fig.add_subplot(section[upper:lower, 23:29])
    '''retrive the data'''
    name = []
    win_rate = []
    colors = []
    for idx in range(1,benchmark_no+1):
        rate = data_before_win_rate[letters[idx] + '2'].value
        if rate>0 :
            name.append(data_before_win_rate[letters[idx] + '1'].value)
            win_rate.append(rate)
            colors.append(color_list[idx])
    index = np.arange(20)
    '''labelling'''
    no = len(win_rate)
    label = [None if win_rate[i]<=2 and win_rate[i-1]<=3  else str(win_rate[i]) for i in range(no)]
    annotates = [str(win_rate[i]) if win_rate[i]<=2 and win_rate[i-1]<=3 else None for i in range(no)]
    wedges,texts = ax[ax_idx+10].pie(win_rate, labels = label, labeldistance=1.2, colors=colors, startangle=90, wedgeprops=dict(ec='w'), textprops=dict(fontsize=8.5))
    plt.setp(texts, ha='center', )
    kw = dict(arrowprops=dict(arrowstyle="-",lw=0.5),  va="center", fontsize=8.5, )
    pre=1
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        if annotates[i]:
            if np.abs(pre-y) > 0.15 or pre==1:
                ax[ax_idx+10].annotate(annotates[i], xy=(x, y), xytext=(1.2*np.sign(x), 1.4*y),horizontalalignment=horizontalalignment, **kw)
            else:
                ax[ax_idx+10].annotate(annotates[i], xy=(x, y), xytext=(1.2*np.sign(x), 1.4*pre-0.18),horizontalalignment=horizontalalignment, **kw)
            pre=y
    ax[ax_idx+10].axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    ax[ax_idx+10].set_title(panel_B[ax_idx], fontsize=12)



    '''and the pie chart'''
    ax[ax_idx+20] = fig.add_subplot(section[upper:lower, 31:37])
    '''retrive the data'''
    name = []
    win_rate = []
    colors = []
    for idx in range(1,benchmark_no+1):
        rate = data_win_rate[letters[idx] + '2'].value
        if rate>0 :
            name.append(data_win_rate[letters[idx] + '1'].value)
            win_rate.append(rate)
            colors.append(color_list[idx])
    index = np.arange(20)
    explode = [0 for i in range(len(name)-1)]+[0.15]  # only "explode" DRL
    '''labelling'''
    no = len(win_rate)
    label = [None if win_rate[i]<=2 and win_rate[i-1]<=3 else str(win_rate[i]) for i in range(no)]
    annotates = [str(win_rate[i]) if win_rate[i]<=2 and win_rate[i-1]<=3 else None for i in range(no)]
    wedges,texts = ax[ax_idx+20].pie(win_rate, explode=explode, labels = label, radius=1, labeldistance=1.2, colors=colors, startangle=90, wedgeprops=dict(ec='w'), textprops=dict(fontsize=8.5))
    plt.setp(texts, ha='center', )
    plt.setp(texts[-1], fontsize=10.5, fontweight='bold')
    plt.setp(wedges[-1], ec='k')
    kw = dict(arrowprops=dict(arrowstyle="-",lw=0.5),  va="center", fontsize=8.5, )
    pre=1
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        if annotates[i]:
            if np.abs(pre-y) > 0.15 or pre==1:
                ax[ax_idx+20].annotate(annotates[i], xy=(x, y), xytext=(1.2*np.sign(x), 1.4*y),horizontalalignment=horizontalalignment, **kw)
            else :
                ax[ax_idx+20].annotate(annotates[i], xy=(x, y), xytext=(1.2*np.sign(x), 1.4*pre-0.18),horizontalalignment=horizontalalignment, **kw)
            pre=y
    ax[ax_idx+20].axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    ax[ax_idx+20].set_title(panel_C[ax_idx], fontsize=12)



fig.subplots_adjust(top=0.95, bottom=0.1, hspace=0.5)
#fig.savefig(sys.path[0]+"/pics/experiment_result.png", dpi=600, bbox_inches='tight')
plt.show()
