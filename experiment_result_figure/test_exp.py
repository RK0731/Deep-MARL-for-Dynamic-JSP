import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
import matplotlib.lines as mlines
import matplotlib.ticker as mtick
import sys
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()


# index to retrive the data
panel_A = ['(a.1) Performance gain in 70% scenario','(b.1) Performance gain in 80% scenario','(c.1) Performance gain in 90% scenario']
panel_B = ['(a.2) Win rate in 70% scenario','(b.2) Win rate in 80% scenario','(c.2) Win rate in 90% scenario']

letters = string.ascii_uppercase
benchmark_no = 20
scenarios = [70,80,90]
runs = 100
color_list = ['#7e1e9c','#15b01a','#0343df','#ff81c0','#653700','#e50000','#95d0fc','#029386','#f97306','#c2b709','#c20078','#00035b','#75bbfd','#929591','#89fe05','#8f1402','#9a0eea','#033500','#06c2ac','#ffffe4','#8af1fe','#d1b26f','#00ffff','#13eac9']
bplot_color_list = ['None' for i in range(benchmark_no-1)]+['#8af1fe']
'''plot'''
# gridspec inside gridspec
fig = plt.figure(figsize=(12,10))
ax = {}
sec_ax = {}
all_scenarios = gridspec.GridSpec(3, 1, figure=fig)


for ax_idx in range(3):
    '''import the data'''
    path = sys.path[0]+'\exp_result_{}.xlsx'.format(scenarios[ax_idx])
    data_sum = load_workbook(path)['sum']
    data_win_rate = load_workbook(path)['win rate']
    '''create the grids'''
    section = gridspec.GridSpecFromSubplotSpec(30, 30, subplot_spec = all_scenarios[ax_idx])
    ax[ax_idx] = fig.add_subplot(section[:, :20])
    # set different range to create broken axis
    ax[ax_idx].set_ylim(bottom=0, top=0.9)
    ax[ax_idx].set_xlim(left = -1, right=benchmark_no)
    '''retrive the data'''
    name = []
    sum = []
    # retrive the data
    for idx in range(benchmark_no+1):
        name.append(data_sum[letters[idx] + '1'].value)
        sum.append([data_sum[letters[idx] + str(i)].value for i in range(2,2+runs)])
    #print(name)
    #print(sum)
    sum = 1 - sum / np.array(sum[0])
    #print(sum)
    name.pop(0) # drop the FIFO
    sum = np.delete(sum,0,axis=0)
    # create the plots
    x_position = np.arange(len(name))
    '''plot the data'''
    bplot = ax[ax_idx].boxplot(sum.transpose(), positions=x_position, showmeans=True, meanline=True, patch_artist=True, notch=True, zorder=3,)
    for patch, c in zip(bplot['boxes'], bplot_color_list):
        patch.set_facecolor(c)
    # ax[ax_idx].violinplot(sum.transpose(), positions=x_position, showmeans=True, )
    # ticks
    ax[ax_idx].set_yticks(np.arange(-0.2,1,0.1))
    ax[ax_idx].yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    ax[ax_idx].set_xticks(x_position)
    ax[ax_idx].set_xticklabels(name)
    ax[ax_idx].set_ylabel('Normalized performance %', fontsize=10)
    #ax[ax_idx].set_yticks(np.arange(0, 1.5, 0.1))
    plt.setp(ax[ax_idx].get_xticklabels(), rotation=30, ha='right', rotation_mode="anchor", fontsize=9)
    plt.setp(ax[ax_idx].get_yticklabels(), fontsize=9)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )
    # lines
    ax[ax_idx].hlines(y=0, xmin=-1, xmax=20, colors='k', linestyles='solid', linewidths=0.5)
    ax[ax_idx].hlines(y=sum[-1].mean(), xmin=-1, xmax=20, colors='r', linestyles='--', linewidths=0.5, zorder=1)
    # label
    ax[ax_idx].set_title(panel_A[ax_idx], fontsize=10)
    '''common legend'''
    if ax_idx == 1:
        ax[100] = fig.add_subplot(section[:, 28:])
        # common legend
        legend_elements = [Patch(facecolor=color_list[i+1],label=name[i]) for i in range(benchmark_no)]
        ax[100].legend(handles=legend_elements, fontsize=9, loc=6)
        ax[100].axis('off')
    '''and the pie chart'''
    ax[ax_idx+10] = fig.add_subplot(section[:, 22:27])
    '''retrive the data'''
    name = []
    win_rate = []
    colors = []
    # retrive the data
    for idx in range(1,benchmark_no+1):
        rate = data_win_rate[letters[idx] + '2'].value
        if rate>0 :
            name.append(data_win_rate[letters[idx] + '1'].value)
            win_rate.append(rate)
            colors.append(color_list[idx])
    index = np.arange(20)
    explode = [0 for i in range(len(name)-1)]+[0.15]  # only "explode" DRL
    ax[ax_idx+10].pie(win_rate, explode=explode, autopct='%1.0f%%',pctdistance=1.2, colors=colors, startangle=90, wedgeprops=dict(ec='w'), textprops=dict(fontsize=8))
    ax[ax_idx+10].axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    ax[ax_idx+10].set_title(panel_B[ax_idx], fontsize=10)


fig.subplots_adjust(top=0.95, bottom=0.1, wspace=0.15, hspace=0.5)
#fig.savefig(sys.path[0]+"/pics/experiment_result.png", dpi=500, bbox_inches='tight')
plt.show()
