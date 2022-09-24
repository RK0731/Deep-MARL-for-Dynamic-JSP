import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
from matplotlib.lines import Line2D
import matplotlib.ticker as mtick
import sys
sys.path
import numpy as np
import string
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()


'''import the data'''
path = sys.path[0]+'\\decision_influence.xlsx'
data_decision_influence = load_workbook(path)['influence']
data_tardy_rate = load_workbook(path)['tardy_rate']
data_tardiness = load_workbook(path)['tardiness']

# index to retrive the data_decision_influence
title_col = 1
name_col = 0
ones_col = 1
twos_col = 2
threes_col = 3
fours_col = 4
aboves_col = 5
panels = []
letters = string.ascii_uppercase
color_list = ['w','#d1ffbd','#39ad48','#137e6d','k']
edge_list =  ['k','#d1ffbd','#39ad48','#137e6d','k']
#color_list = ['w','#d0fefe','#a2cffe','#0165fc','k']
#edge_list =  ['k','#d0fefe','#a2cffe','#0165fc','k']
hatch_list = ['//']+[None for i in range(4)]
legend_list = ['passive decision','two jobs','three jobs', 'four jobs', 'five jobs and above']
scenario_no = 5

# figure and subplots
ax={}
fig = plt.figure(figsize=(11,8))

#ax_legend = fig.add_subplot(111, visible=False)

''' plot the decision influence'''
for ax_idx in range(3): # draw subplot for each scenario
    ax[ax_idx] = fig.add_subplot(3,4,ax_idx+1)
    title = data_decision_influence[letters[title_col] + '1'].value
    name = []
    ones = []
    twos = []
    threes = []
    fours = []
    aboves = []
    winning_rate = []
    # retrive the data_decision_influence
    for idx in range(3,3+scenario_no):
        name.append(data_decision_influence[letters[name_col] + str(idx)].value)
        ones.append(data_decision_influence[letters[ones_col] + str(idx)].value)
        twos.append(data_decision_influence[letters[twos_col] + str(idx)].value)
        threes.append(data_decision_influence[letters[threes_col] + str(idx)].value)
        fours.append(data_decision_influence[letters[fours_col] + str(idx)].value)
        aboves.append(data_decision_influence[letters[aboves_col] + str(idx)].value)
    bars = [ones,twos,threes,fours,aboves]
    bottom = np.array([np.zeros(scenario_no),ones,twos,threes,fours])
    #print(bars)
    #print(bottom)
    bottom = np.cumsum(bottom,axis=0)
    # x position of bars
    x_position = np.arange(scenario_no)
    # limit
    y_range = 1.0
    ax[ax_idx].set_ylim(0,y_range)
    # plot the bars
    for i in range(scenario_no):
        ax[ax_idx].bar(x_position, bars[i], 0.4,  bottom=bottom[i], color=color_list[i], hatch=hatch_list[i], edgecolor=edge_list[i], align = 'center', zorder=3,)
    # ticks
    ax[ax_idx].set_xticks(np.arange(len(name)))
    ax[ax_idx].set_xticklabels(name)
    ax[ax_idx].set_yticks(np.arange(0, 1.05, 0.1))
    ax[ax_idx].yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    plt.setp(ax[ax_idx].get_xticklabels(), fontsize=9)
    plt.setp(ax[0].get_yticklabels(), visible=True)
    # labels
    ax[ax_idx].set_title(title, fontsize=10)
    ax[0].set_ylabel('(1) Ratio of decisions %', fontsize=10)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )
    # cols
    title_col += 5
    ones_col += 5
    twos_col += 5
    threes_col += 5
    fours_col += 5
    aboves_col += 5

''' plot the tardy rate'''
first_col = 1
for ax_idx in range(5,8): # draw subplot for each scenario
    ax[ax_idx] = fig.add_subplot(3,4,ax_idx)
    title = data_tardy_rate[letters[title_col] + '1'].value
    t_rate = []
    x_position = np.arange(scenario_no)
    for col in range(first_col, first_col+scenario_no):
        t_rate.append([data_tardy_rate[letters[col] + str(idx)].value for idx in range(3,53)])
    #print(t_rate)
    ax[ax_idx].boxplot(t_rate, positions=x_position, showmeans=True, meanline=True, notch=True, zorder=3,)
    # ticks
    ax[ax_idx].set_xticks(np.arange(len(name)))
    ax[ax_idx].set_xticklabels(name)
    ax[ax_idx].set_yticks(np.arange(0, 1.05, 0.1))
    ax[ax_idx].yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1,symbol=None))
    ax[5].set_ylabel('(2) Tardy Rate %', fontsize=10)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )
    # cols
    first_col += 5

''' plot the tardiness'''
first_col = 1
for ax_idx in range(9,12): # draw subplot for each scenario
    ax[ax_idx] = fig.add_subplot(3,4,ax_idx)
    title = data_tardiness[letters[title_col] + '1'].value
    t_rate = []
    x_position = np.arange(scenario_no)
    for col in range(first_col, first_col+scenario_no):
        t_rate.append([data_tardiness[letters[col] + str(idx)].value for idx in range(3,53)])
    #print(t_rate)
    ax[ax_idx].boxplot(t_rate, positions=x_position, showmeans=True, meanline=True, notch=True, zorder=3,)
    # ticks
    ax[ax_idx].set_xticks(np.arange(len(name)))
    ax[ax_idx].set_xticklabels(name)
    ax[ax_idx].set_yticks(np.arange(0, 310, 50))
    ax[ax_idx].set_ylim(bottom=0)
    ax[9].set_ylabel('(3) Average Tardiness', fontsize=10)
    # grid
    ax[ax_idx].grid(axis='y', which='major', alpha=0.5, zorder=0, )
    # cols
    first_col += 5


'''common legends'''
ax_legend_0 = fig.add_subplot(3,4,4)
color_list.reverse()
edge_list.reverse()
hatch_list.reverse()
legend_list.reverse()
legend_elements = [Patch(facecolor=color_list[i], edgecolor=edge_list[i], hatch=hatch_list[i], label=legend_list[i]) for i in range(5)]
ax_legend_0.legend(handles=legend_elements, fontsize=10, loc=7)
ax_legend_0.axis('off')
# boxplot legend
ax_legend_1 = fig.add_subplot(2,4,8)
dummy = np.array([0.5,2,2,2,3,4,5,7,8,9,6,7,4,5,6,3,5,1,2,4,13])
ax_legend_1.set_xlim(0.9,1.8)
ax_legend_1.set_ylim(-6,13.5)
ax_legend_1.set_xticks(np.arange(1,2,0.5))
ax_legend_1.boxplot(dummy,showmeans=True, meanline=True, notch=True)
ax_legend_1.text(1.1,dummy.mean()-0.15,'mean',color='g')
ax_legend_1.text(1.1,np.median(dummy)-0.55,'median',color='tab:orange')
ax_legend_1.text(1.1,np.percentile(dummy,75),'75th percentile', verticalalignment='center')
ax_legend_1.text(1.1,np.percentile(dummy,25),'25th percentile', verticalalignment='center')
ax_legend_1.text(1.1,9,'Maximum', verticalalignment='center')
ax_legend_1.text(1.1,dummy.min(),'Minimum', verticalalignment='center')
ax_legend_1.text(1.1,dummy.max(),'Outlier', verticalalignment='center')
ax_legend_1.axis('off')
ax[10].set_xlabel('Number of Machines', fontsize=10)

fig.subplots_adjust(top=0.9, bottom=0.1, right=0.9, wspace=0.25, hspace=0.15)
#fig.savefig(sys.path[0]+"/pics/Thesis_influence_tardy_rate.png", dpi=600, bbox_inches='tight')
plt.show()
